import json
import os
import sys
import shutil
import time
import re
import argparse
from datetime import datetime
import subprocess

import pandas

# import readline
from prompt_toolkit import prompt
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from bs4 import BeautifulSoup
from typing import Any, Iterable

from pandas import DataFrame
from tqdm import tqdm

from lib import helper_functions
from lib import selenium_driver_functions
from lib.testmo_functions import download_fields_and_csv, login_to_testmo
from lib.testmo_access import (
    testmo,
    testmo_collect,
    testmo_project_runs_request,
    testmo_project_run_results_request,
    testmo_result_colors_by_code,
    testmo_result_status_names,
    testmo_result_status,
    testmo_config,
    testmo_result_colors,
    testmo_config_file,
)
from lib.rest_request import RestRequest
from lib.json_cursor import JsonCursor
from lib.dict_util import dict_from
from lib.pandas_dataframe_filter import FilterBase, FilterRegistry
from lib.file_access import restrict_to_owner

from schema import (
    testmo_project_info_reply,
    testmo_project_run_reply,
    testmo_project_run_result,
)


testmo_request: RestRequest = None
testmo_projects_request: RestRequest = None
cell_alignmnent_left = Alignment(horizontal="left")
bold_font = Font(bold=True)
std_fields = ["Case ID", "Case", "Folder", "State", "Status (latest)"]


# ------- Filter Definitions -------
class RemoveDeletedFilter(FilterBase):
    @classmethod
    def description(cls) -> str:
        return "not deleted"

    @classmethod
    def name(cls):
        return None

    @classmethod
    def apply_to(cls, dataframe: DataFrame) -> DataFrame:
        """Exclude if folder name begins with '(Deleted)'"""
        result = dataframe[~dataframe["Folder"].str.startswith("(Deleted)")]
        return result


class NoFilter(FilterBase):
    @classmethod
    def description(cls) -> str:
        return "No filter"

    @classmethod
    def name(cls):
        return "none"

    @classmethod
    def apply_to(cls, dataframe: DataFrame) -> DataFrame:
        return dataframe


class ActiveFilter(FilterBase):
    @classmethod
    def description(cls) -> str:
        return "State = Active"

    @classmethod
    def name(cls):
        return "active"

    @classmethod
    def apply_to(cls, dataframe: DataFrame) -> DataFrame:
        """Include if state is 'Active'"""
        result = dataframe[dataframe["State"] == "Active"]
        return result


class SafetyFilter(FilterBase):
    @classmethod
    def description(cls) -> str:
        return "Safety = Yes"

    @classmethod
    def name(cls):
        return "safety"

    @classmethod
    def apply_to(cls, dataframe: DataFrame) -> DataFrame:
        """Include if safety is present and Yes"""
        if "Safety" not in dataframe.columns:
            return dataframe

        result = dataframe[dataframe["Safety"].isin({"Yes"})]
        return result


class RemoveRetiredAndRejected(FilterBase):
    @classmethod
    def description(cls) -> str:
        return "State \u2260 Retired,Rejected"

    @classmethod
    def name(cls):
        return "not_retired_or_rejected"

    @classmethod
    def apply_to(cls, dataframe: DataFrame) -> DataFrame:
        """Include if state is 'Active'"""
        result = dataframe[~dataframe["State"].isin({"Retired", "Rejected"})]
        return result


# ------- End Filter Definitions -------


def get_filters(filter_description: str):
    def _get_filter_name(name):
        try:
            tmp = int(name)
        except Exception:
            tmp = None

        available_filters = [_ for _ in all_filters]
        if tmp is not None and tmp >= 1 and tmp <= len(available_filters):
            filter_name = available_filters[tmp - 1]
        else:
            filter_name = name

        _result = FilterRegistry.get_filter(filter_name)
        return _result

    def _normalize_names(_line):
        _filters = split_input(_line)
        _result = {f: _get_filter_name(f) for f in _filters}
        return _result

    all_filters = FilterRegistry.all_filters()
    selected_filters = _normalize_names(filter_description)
    return selected_filters


def open_with_default_app(filename):
    """
    Open a file with the default application associated with its extension,
    supporting Windows, macOS, and Linux.
    If unsupported, prints a notification.
    """
    try:
        if sys.platform.startswith("win"):
            # Windows
            os.startfile(filename)
        elif sys.platform == "darwin":
            # macOS
            subprocess.run(["open", filename], check=True)
        elif sys.platform.startswith("linux"):
            # Most desktop Linux distros
            subprocess.run(["xdg-open", filename], check=True)
        else:
            print(f"No known method to open files for OS: {sys.platform}")
    except Exception as e:
        print(f"Could not open file '{filename}': {e}")


def is_method_overridden(obj, method_name, base_class):
    """
    Returns True if `method_name` is overridden in obj's class (or any subclass up to base_class),
    compared to `base_class`. Otherwise, False.
    """
    # Get the method from the instance's class
    sub_method = getattr(obj.__class__, method_name, None)
    # Get the method from the base class
    base_method = getattr(base_class, method_name, None)

    if sub_method is None or base_method is None:
        return False  # Method doesn't exist somewhere

    # Unwrap if necessary
    sub_func = getattr(sub_method, "__func__", sub_method)
    base_func = getattr(base_method, "__func__", base_method)
    return sub_func is not base_func


def request_setup(testmo_url: str, testmo_token: str):
    """Prepare RestRequests"""
    global testmo_request, testmo_projects_request
    testmo_request = (
        testmo
        if testmo is not None and testmo_token is None
        else RestRequest(
            base_url=testmo_url,
            headers={
                "accept": "application/json",
                "Authorization": f"Bearer {testmo_token}",
            },
        )
    )
    if testmo_url is not None:
        testmo_request.modify(base_url=testmo_url)

    testmo_projects_request = testmo_request.copy(endpoint="projects")


def keep_text(text):
    return text


def display_choices(
    choices: Iterable[Any],
    text_key: str = None,
    index_key: str = None,
    limiter=keep_text,
):
    """
    Display a list of choices (as dicts) in as many columns as will fit in the terminal.
    Expects a list of dicts with keys 'index' (for numbering) and 'name' (for text).
    """
    if not choices:
        print("No choices available.")
        return

    # Get terminal width
    try:
        columns = shutil.get_terminal_size().columns
    except Exception:
        columns = 80  # Fallback

    if columns == 80:  # ensure min. 2 columns
        columns = 120

    # Build the display strings using 'index' and 'name'
    if text_key is None:
        indexed_choices = [
            f"{str(i).rjust(4)}. {limiter(item)}" for i, item in enumerate(choices, 1)
        ]
    elif index_key is not None:
        indexed_choices = [
            f"{str(item[index_key]).rjust(4)}. {limiter(item[text_key])}"
            for item in choices
        ]
    else:
        indexed_choices = [
            f"{str(i).rjust(4)}. {limiter(item[text_key])}"
            for i, item in enumerate(choices, 1)
        ]

    # Find the max length of each entry
    max_len = max(len(item) for item in indexed_choices)
    col_width = max_len + 4  # padding between columns

    # Figure out how many columns can fit
    num_columns = max(1, columns // col_width)
    num_rows = (len(indexed_choices) + num_columns - 1) // num_columns

    # Arrange the choices in a grid: rows × columns
    grid = []
    for row in range(num_rows):
        line = []
        for col in range(num_columns):
            idx = col * num_rows + row
            if idx < len(indexed_choices):
                line.append(indexed_choices[idx].ljust(col_width))
        grid.append("".join(line).rstrip())

    # Print the grid
    for line in grid:
        print(line)


def ask_for_input(prompt_str: str, preset: str = ""):
    """
    Ask for user input.

    Args:
            prompt (str): the user prompt
            preset (str): preset value for the user input

    Returns:
            (str): the user input
    """
    try:
        user_input = prompt(prompt_str + ": ", default=preset)
        return user_input
    except Exception as e:
        # print(f"(advanced prompt unavailable: {e})")
        # Show preset in brackets as a fallback
        default_prompt = f"{prompt_str} [{preset}]: " if preset else f"{prompt_str}: "
        user_input = input(default_prompt)
        return user_input if user_input else preset
        pass


def load_xpaths():
    json_path = os.path.join("json", "xpaths.json")
    result = helper_functions.load_json(json_path)
    return result


def get_project_mapping():
    projects = testmo_collect(testmo_projects_request)
    result = [
        {"id": str(p["id"]), "name": p["name"]}
        for p in sorted(projects, key=lambda e: e["name"])
    ]
    return result


def get_project_info(project_name: str) -> testmo_project_info_reply.Result:
    testmo_projects = testmo_collect(testmo_projects_request)
    project_info = JsonCursor(testmo_projects).search(
        lambda e: isinstance(e.node, dict)
        and (e.node.get("name") == project_name or e.node.get("id") == project_name)
    )
    return (
        testmo_project_info_reply.result_from_data(project_info.node)
        if project_info is not None
        else None
    )


def autofit_column_widths(ws, min_column_width=18):
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        adjusted_width = max(max_length + 2, min_column_width)
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width


def update_statistics(stats, test_result: str):
    match test_result:
        case "Passed":
            stats["passed"] += 1
        case "Failed":
            stats["failed"] += 1
        case _:
            stats["other"] += 1


def testrun_is_active(test_run: testmo_project_run_reply.Model) -> bool:
    result = test_run.is_started and not test_run.is_closed
    return result


def table_to_sheet(
    project_info: testmo_project_info_reply.Result,
    csv_path: str,
    additional_columns,
    case_filter,
    number_of_runs,
    event_handler: "ApplicationEventHandler",
):
    """
    Main conversion repository table => Excel sheet
    """

    def _create_sheet_columns(_sheet, _columns):
        _latest_state_column = -1
        for i, column in enumerate(_columns, table_start[1]):
            cell = _sheet.cell(row=table_start[0], column=i)
            cell.value = column
            cell.font = bold_font
            if column == "Status (latest)":
                _latest_state_column = i
        return _latest_state_column

    def _create_testcase_rows(_sheet, _table, _column_names) -> set[int]:
        nonlocal row_by_case_id
        tracing_required = set()
        for index, _row in _table.iterrows():
            # for all rows ...
            row = table_start[0] + 1 + index  # first row are column names
            case_id = _table.iloc[index]["Case ID"]
            row_by_case_id[case_id] = row
            for col, entry in enumerate(_column_names, table_start[1]):
                # ... process each column
                cell = _sheet.cell(row=row, column=col)
                try:
                    cell.value = _table.iloc[index][entry]
                except KeyError:
                    cell.value = "--"
                cell.alignment = cell_alignmnent_left

                if entry != "Status (latest)":
                    continue

                # special case for latest status: coloring, statistics and tracing of blocked, skipped and retest
                try:
                    case_color = testmo_result_colors.__dict__[cell.value]
                    if case_color is not None:
                        cell.fill = PatternFill(
                            start_color=case_color,
                            end_color=case_color,
                            fill_type="solid",
                        )
                except Exception:
                    pass

                match cell.value:
                    case "Blocked" | "Skipped" | "Retest":
                        tracing_required.add(case_id)
                    case _:
                        update_statistics(stats, cell.value)
                    # case "Passed":
                    # 	stats['passed'] += 1
                    # case "Failed":
                    # 	stats['failed'] += 1
                    # case "Blocked" | "Skipped" | "Retest":
                    # 	tracing_required.add(case_id)
                    # case _:
                    # 	stats['other'] += 1

        return tracing_required

    def _insert_test_run_results(
        _sheet,
        _table,
        _test_run: testmo_project_run_reply.Model,
        _run_index: int,
        _column_offset: int,
        _traced_cases_in_runs,
    ):
        nonlocal row_by_case_id, stats, traced_cases
        # ------ load test data and prepare insert ------
        test_cases = testmo_collect(
            testmo_project_run_results_request(
                _test_run.id, base_request=testmo_request
            ),
            convert_to=testmo_project_run_result.from_data,
        )
        run_column_index = runs_start[1] + _run_index - _column_offset
        run_title_cell = _sheet.cell(row=table_start[0], column=run_column_index)
        run_title_cell.value = _test_run.name
        run_title_cell.font = bold_font
        # last_change = {}

        # ------ insert test results ------
        for case_index, case in enumerate(test_cases):
            if not case.is_latest:
                continue
            case_id = case.case_id
            # record case if it's a traced case
            if case_id in traced_cases:
                cases = _traced_cases_in_runs.get(case_id, [])
                cases.insert(0, case)
                _traced_cases_in_runs[case_id] = cases

            # obtain test position in sheet
            case_index_in_sheet = row_by_case_id.get(case_id)
            if case_index_in_sheet is None:
                event_handler.case_id_not_in_project_map(locals())
                continue
            else:
                event_handler.adding_case_to_sheet(locals())

            # approximate last test result
            case_status = case.status_id
            # case_created = datetime.strptime(case.created_at, '%Y-%m-%dT%H:%M:%S.%f%z')
            # if case_id in last_change and last_change[case_id] > case_created:
            # 	continue    # use newest result
            # last_change[case_id] = case_created

            # create excel cell for test result
            case_text = testmo_result_status_names[case_status]
            case_color = testmo_result_colors_by_code[case_status]
            cell = _sheet.cell(row=case_index_in_sheet, column=run_column_index)
            cell.value = case_text
            if case_color is not None:
                cell.fill = PatternFill(
                    start_color=case_color, end_color=case_color, fill_type="solid"
                )
        pass

    def _percentage(_stats, field):
        total = _stats["total"]
        if total == 0:
            return "--"
        percent = 100.0 * _stats[field] / total
        return "%.2f" % percent

    def _filter_description() -> str:
        filters = get_filters(case_filter)
        # filter_class = FilterRegistry.get_filter(case_filter)
        if filters is None or len(filters) == 0:
            return "(no filter)"
        else:
            return f'({" | ".join([f.description() for f in filters.values()])})'

    def _insert_header(_sheet, _stats):
        # Insert header
        logo = Image("Files/murr_logo.png")
        logo.width = int(logo.width / 2)
        logo.height = int(logo.height / 2)
        _sheet.row_dimensions[1].height = logo.height
        _sheet.column_dimensions["A"].width = 30
        _sheet.add_image(logo, "A1")
        _sheet.merge_cells("B1:D1")
        filter_text = _filter_description()
        _sheet["B1"] = f"Test Overview {project_info.name} {filter_text}"
        _sheet["B1"].font = Font(name="Calibri", size=20, bold=True, color="000000")
        _sheet["B1"].alignment = Alignment(horizontal="center", vertical="center")

    def _insert_statistics(_sheet, _stats):
        # Insert Stats
        start_row = 2
        start_column = 1
        for i, column in enumerate(_stats, start_row):
            kcell = _sheet.cell(row=i, column=start_column)
            kcell.value = column
            kcell.font = bold_font
            vcell = _sheet.cell(row=i, column=start_column + 1)
            if i == start_column:
                vcell.value = _stats[column]
            else:
                vcell.value = f"{_stats[column]} / {_percentage(_stats, column)}%"
            vcell.alignment = cell_alignmnent_left
        pass

    def _trace_inconclusive_entries(
        _sheet, _stats, _traced_cases, _traced_cases_in_runs
    ):
        nonlocal row_by_case_id, latest_state_column

        traced_states = {
            testmo_result_status.Blocked,
            testmo_result_status.Skipped,
            testmo_result_status.Retest,
        }
        for _case_id in _traced_cases:
            _cases = _traced_cases_in_runs.get(_case_id)
            _case_state = None
            if _cases is not None:
                # continue
                for _case in _cases:
                    if _case.status_id in traced_states:
                        continue
                    _case_state = _case.status_id
                    break

                update_statistics(_stats, testmo_result_status_names.get(_case_state))

            _row = row_by_case_id.get(_case_id)
            if _row is None:
                continue

            _cell = _sheet.cell(row=_row, column=latest_state_column)
            if _case_state is None:
                _cell.value = None
                _cell.fill = PatternFill()
            else:
                _cell.value = testmo_result_status_names[_case_state]
                case_color = testmo_result_colors.__dict__[_cell.value]
                if case_color is not None:
                    _cell.fill = PatternFill(
                        start_color=case_color, end_color=case_color, fill_type="solid"
                    )
                else:
                    _cell.fill = PatternFill()
        pass

    ############### MAIN FUNCTION BODY ####################
    stats = dict_from(total=0, passed=0, failed=0, other=0, not_in_project=0)

    # ------ obtain repository csv ------
    event_handler.csv_read_started(locals())
    table_raw = pandas.read_csv(csv_path)
    table_cleared = [
        RemoveDeletedFilter.apply_to(table_raw)
    ]  # deleted folders are always removed
    case_filter = event_handler.csv_read_complete(locals())
    table = table_cleared[0].sort_values(by=["Folder", "Case"]).reset_index(drop=True)

    # ------ prepare excel sheet ------
    event_handler.sheet_setup_started(locals())
    excel_workbook = openpyxl.Workbook()
    sheet = excel_workbook.active
    sheet.title = f"{project_info.name} Overview"
    sheet_column_names = [*std_fields, *additional_columns]

    table_start = (9, 1)
    runs_start = (table_start[0], table_start[1] + len(sheet_column_names))
    row_by_case_id = {}
    latest_state_column = _create_sheet_columns(sheet, sheet_column_names)
    traced_cases = _create_testcase_rows(sheet, table, sheet_column_names)
    event_handler.sheet_setup_complete(locals())

    # ------ collect test run data ------
    event_handler.collecting_test_runs(locals())
    test_runs = testmo_collect(
        testmo_project_runs_request(project_info.id, base_request=testmo_request),
        convert_to=testmo_project_run_reply.from_data,
    )
    event_handler.test_runs_collected(locals())

    # number_of_runs == 0 -> active runs, -1 -> all runs, > 0 -> last N runs
    total_run_count = len(test_runs)
    skipped_runs = total_run_count - number_of_runs
    offset = 0
    if number_of_runs >= total_run_count:
        number_of_runs = -1
    elif number_of_runs > 0:
        offset = skipped_runs

    # ------ process test run data ------
    traced_cases_in_runs = {}
    effective_run_index = -1
    for run_index, test_run in enumerate(sorted(test_runs, key=lambda r: r.created_at)):
        if number_of_runs == 0:  # only active runs
            if testrun_is_active(test_run):
                effective_run_index += 1
            else:
                continue
        elif number_of_runs > 0:  # given number of runs
            if run_index < skipped_runs:
                continue
            else:
                effective_run_index = run_index
        else:  # all runs
            effective_run_index = run_index

        event_handler.test_run_processing_started(locals())
        _insert_test_run_results(
            sheet, table, test_run, effective_run_index, offset, traced_cases_in_runs
        )
        event_handler.test_run_processing_complete(locals())

    # ------ trace last state for blocked and skipped last states ------
    event_handler.tracing_inconclusive_entries(locals())
    _trace_inconclusive_entries(sheet, stats, traced_cases, traced_cases_in_runs)

    # ------ insert statistics ------
    event_handler.inserting_statistics(locals())
    autofit_column_widths(sheet)
    _insert_header(sheet, stats)
    _insert_statistics(sheet, stats)

    # ------ save and launch viewer ------
    event_handler.saving_sheet(locals())
    excel_filename = os.path.join("Files", f"{project_info.name}.xlsx")
    excel_workbook.save(excel_filename)
    event_handler.launching_viewer(locals())
    open_with_default_app(excel_filename)


def extract_field_names(html):
    soup = BeautifulSoup(html, "html.parser")
    field_names = []

    # Find all relevant td tags (those with the target class)
    for td in soup.find_all("td", class_="table__field__avatar-text"):
        # The actual field name is the div after the .avatar div
        avatar_div = td.find("div", class_="avatar")
        if avatar_div:
            # Find the next div after avatar_div
            next_div = avatar_div.find_next_sibling("div")
            if next_div and next_div.text.strip():
                field_names.append(next_div.text.strip())
    return field_names


def split_input(line):
    pattern = r'"([^"]*)"|\'([^\']*)\'|([^,\s]+)'

    entries = []
    for match in re.findall(pattern, line):
        # Each match returns a tuple with only one non-empty entry
        token = next(filter(None, match))
        entries.append(token)

    return entries


class ApplicationEventHandler:
    """All available process events base class. Attention: only concrete handlers will get called."""

    def application_setup(self, kwargs):
        pass

    def application_started(self, kwargs):
        pass

    def project_does_not_exist(self, kwargs):
        pass

    def project_started(self, kwargs):
        pass

    def adding_case_to_sheet(self, kwargs):
        pass

    def case_id_not_in_project_map(self, kwargs):
        pass

    def selenium_setup(self, kwargs):
        pass

    def testmo_gui_login(self, kwargs):
        pass

    def csv_download_started(self, kwargs):
        pass

    def csv_download_fields_received(self, kwargs):
        pass

    def csv_download_complete(self, kwargs):
        pass

    def csv_read_started(self, kwargs):
        pass

    def csv_read_complete(self, kwargs):
        pass

    def sheet_setup_started(self, kwargs):
        pass

    def sheet_setup_complete(self, kwargs):
        pass

    def collecting_test_runs(self, kwargs):
        pass

    def test_runs_collected(self, kwargs):
        pass

    def test_run_processing_started(self, kwargs):
        pass

    def test_run_processing_complete(self, kwargs):
        pass

    def inserting_statistics(self, kwargs):
        pass

    def tracing_inconclusive_entries(self, kwargs):
        pass

    def saving_sheet(self, kwargs):
        pass

    def launching_viewer(self, kwargs):
        pass

    def processing_complete(self, kwargs):
        pass


class HandlerChain(ApplicationEventHandler):
    class Abort(Exception):
        pass

    """
	Create a chain of ApplicationEventHandlers, dispatching method calls to them in order.
	*** ATTENTION: overrides __getattribute__: only HandlerChain.Abort and methods of the handlers can be addressed ***
	"""

    def __init__(self, *handlers):
        self.handlers = handlers

    def __getattribute__(self, name: str):
        def _call_handlers(kwargs):
            handlers = object.__getattribute__(self, "handlers")
            result = None
            for handler in handlers:
                if not is_method_overridden(handler, name, ApplicationEventHandler):
                    continue

                method = getattr(handler, name)
                try:
                    call_result = method(kwargs)
                    if call_result is not None:
                        result = call_result
                except HandlerChain.Abort:
                    break
            return result

        if name == "Abort":
            return object.__getattribute__(self, "Abort")

        return _call_handlers


class StdEventHandler(ApplicationEventHandler):
    """mode-independent processing"""

    def project_does_not_exist(self, kwargs):
        project_name = kwargs["project"]
        print(f"Error: project '{project_name}' not found -- exiting.")
        sys.exit(1)

    def case_id_not_in_project_map(self, kwargs):
        kwargs["stats"]["not_in_project"] += 1
        pass


class ApplicationSetup(ApplicationEventHandler):
    """Setup strategy -- selects mode strategy"""

    def application_setup(self, kwargs):
        args = kwargs["args"]
        request_setup(args["testmo_url"], args["testmo_token"])

        required_fields = ["user", "password", "token"]
        for field in required_fields:
            arg_field = f"testmo_{field}"
            if args[arg_field] is None:
                if testmo_config is None:
                    raise ValueError(
                        f"{arg_field} is required but neither provided in config not as parameter"
                    )
                else:
                    args[arg_field] = testmo_config[field]

        kwargs["event_handler"][0] = HandlerChain(InterActiveMode(), StatusMessage())
        pass


class InterActiveMode(StdEventHandler):
    """Processing strategy for interactive mode"""

    def application_started(self, kwargs):
        # validate and format arguments / ask for project if omitted

        def _get_project_name(name):
            for entry in project_mapping:
                if entry["id"] == name:
                    return entry["name"]

            try:
                tmp = int(name)
            except Exception:
                return name

            raise ValueError(f"no such project ID: {name}")

        def _normalize_project(line):
            entries = split_input(line)
            result = [_get_project_name(e) for e in entries]
            return result

        args = kwargs["args"]
        project = args["project_name"]
        project_mapping = get_project_mapping()

        if project is None:
            display_choices(project_mapping, index_key="id", text_key="name")
            project = ask_for_input("Enter one or more projects")

        args["project_name"] = _normalize_project(project)
        pass

    def csv_download_fields_received(self, kwargs):
        """gather available fields in csv, let the user chose and handover selection to download co-routine"""

        def _get_field_name(name):
            try:
                tmp = int(name)
            except Exception:
                _result = name
                tmp = None

            if tmp is not None and tmp >= 1 and tmp < len(fields):
                _result = fields[tmp - 1]
            else:
                _result = name

            if _result not in xpaths_export_entries:
                print(f"Warning: no xpath listed for entry '{_result}'")
            return _result

        def _normalize_fields(line):
            entries = split_input(line)
            result = [_get_field_name(e) for e in entries]
            return result

        xpaths = kwargs["download_args"]["xpaths_dict"]
        xpaths_export_entries = xpaths["TESTMO_REPOSITORY_PAGE"]["EXPORT_CSV_WINDOW"]
        generator = kwargs["csv_generator"]
        if len(kwargs["testmo_additional_fields"]) > 0:
            additional_fields = kwargs["testmo_additional_fields"]
        else:
            fields = [f for f in kwargs["fields"] if f not in std_fields]
            print()
            display_choices(fields)
            choice = ask_for_input("Choose additional fields")
            additional_fields = _normalize_fields(choice)
            kwargs["testmo_additional_fields"].extend(additional_fields)

        all_fields = std_fields[:]
        all_fields.extend(additional_fields)
        result = generator.send(all_fields)
        return result

    def csv_read_started(self, kwargs):
        pass

    def csv_read_complete(self, kwargs):
        table = kwargs["table_cleared"][0]
        filter_names = kwargs["case_filter"]

        display_choices(FilterRegistry.all_filters())
        filter_names = ask_for_input(
            "Choose filters to apply", preset=filter_names + " "
        )
        filters = get_filters(filter_names)

        for filter_name, frame_filter in filters.items():
            if frame_filter is not None:
                table = frame_filter.apply_to(table)

        kwargs["table_cleared"][0] = table
        kwargs["stats"]["total"] = len(table)
        return filter_names


class StatusMessage(ApplicationEventHandler):
    def __init__(self):
        self.progress_bar = None
        self.case_count = 0
        self.current_case = 0

    def csv_download_started(self, kwargs):
        print("Scraping project info from Testmo...")

    def csv_read_complete(self, kwargs):
        filter_names = kwargs["case_filter"]
        filters = get_filters(filter_names)

        for filter_name, frame_filter in filters.items():
            if frame_filter is None:
                print(
                    f"Warning: filter '{filter_name}' is unknown -- filter not applied."
                )

    def csv_download_complete(self, kwargs):
        print("Project info from Testmo received.")

    def project_started(self, kwargs):
        print(f'\nProcessing project {kwargs["project_name"]}')

    def collecting_test_runs(self, kwargs):
        print("Collecting test runs... ", end="")

    def test_runs_collected(self, kwargs):
        print(f'{len(kwargs["test_runs"])} runs found.')
        pass

    def test_run_processing_started(self, kwargs):
        self.case_count = kwargs["test_run"].completed_count
        self.current_case = 0
        self.progress_bar = tqdm(
            total=self.case_count,
            desc=f"Run {kwargs['run_index']+1} ({kwargs['test_run'].id})",
            bar_format="{l_bar}{bar}",
        )
        pass

    def test_run_processing_complete(self, kwargs):
        self.progress_bar.update(self.case_count)
        self.progress_bar.close()

    def adding_case_to_sheet(self, kwargs):
        self.current_case += 1
        self.progress_bar.update(1)
        pass


if __name__ == "__main__":

    def selenium_download_csv(
        event_handler: ApplicationEventHandler,
        project_name,
        testmo_gui_url,
        testmo_user,
        testmo_password,
        testmo_url,
        testmo_token,
        testmo_additional_fields,
        project_info,
        headless=True,
    ) -> str:
        # setup xpaths and table column keys
        testmo_xpaths = load_xpaths()
        project_download_path = os.path.join(os.getcwd(), "Files")
        testmo_csv_keys = std_fields[:]
        for k in testmo_csv_keys:
            if k in testmo_additional_fields:
                del testmo_additional_fields[k]
        testmo_csv_keys.extend(testmo_additional_fields)

        # prepare selenium
        event_handler.selenium_setup(locals())
        selenium_driver = selenium_driver_functions.setup_driver(
            project_download_path, headless=headless
        )

        # login to testmo
        login_args = dict_from(
            xpath_dict=testmo_xpaths,
            testmo_email=testmo_user,
            testmo_password=testmo_password,
        )
        event_handler.testmo_gui_login(locals())
        login_to_testmo(selenium_driver, **login_args)
        time.sleep(2)

        # carry out csv download
        download_args = dict_from(
            project_id=project_info.id,
            project_name=project_name,
            testmo_gui_url=testmo_gui_url,
            download_path=project_download_path,
            xpaths_dict=testmo_xpaths,
            overwrite_existing_file=True,
        )
        event_handler.csv_download_started(locals())
        # obtain download co-routine ...
        csv_generator = download_fields_and_csv(selenium_driver, **download_args)
        # ... and start it
        table_html = next(csv_generator)
        fields = extract_field_names(table_html)
        download_result = event_handler.csv_download_fields_received(locals())
        return download_result

    def process_project(
        event_handler: ApplicationEventHandler,
        project_name,
        testmo_gui_url,
        testmo_user,
        testmo_password,
        testmo_url,
        testmo_token,
        testmo_additional_fields,
        case_filter,
        number_of_runs,
    ):
        request_setup(testmo_url, testmo_token)

        # check project
        project_file_name = (
            "".join(c for c in project_name if c.isalpha() or c.isdigit() or c == " ")
            .rstrip()
            .replace(" ", "")
        )
        download_result = f"Files/{project_file_name}.csv"
        should_download_csv = [True]

        project_info = get_project_info(project_name)
        if project_info is None:
            event_handler.project_does_not_exist(locals())
        else:
            event_handler.project_started(locals())

        # download project repository
        if should_download_csv[0]:
            download_result = selenium_download_csv(
                event_handler,
                project_name,
                testmo_gui_url,
                testmo_user,
                testmo_password,
                testmo_url,
                testmo_token,
                testmo_additional_fields,
                project_info,
                headless=True,
            )

        # create report
        table_to_sheet(
            project_info,
            download_result,
            testmo_additional_fields,
            case_filter,
            number_of_runs,
            event_handler,
        )
        event_handler.processing_complete(locals())
        pass

    def startup_check():
        if os.path.exists(testmo_config_file):
            return

        print(
            f'\nConfiguration file "{testmo_config_file}" not found - Setup required.'
        )
        print("Please provide your Testmo credentials:")
        fields = {
            "user": "Your Testmo Username",
            "password": "Your Testmo Password",
            "token": "Your Testmo API Token",
        }

        template = {
            "gui_url": "https://murrelektronik.testmo.net",
            "url": "https://murrelektronik.testmo.net/api/v1",
        }
        for field in fields:
            answer = ask_for_input(f"{fields[field]}").strip()
            if len(answer) == 0:
                print("Configuration aborted.")
                sys.exit(1)
            template[field] = answer

        with open(testmo_config_file, "w") as f:
            json.dump(template, f, indent=2)

        restrict_to_owner(testmo_config_file)
        print("Config created. Restart the program.\n")
        sys.exit(0)

    def main():
        startup_check()

        parser = argparse.ArgumentParser(
            description="Create overview Excelsheets for Testmo-Projects."
        )
        parser.add_argument(
            "-pn",
            "--project_name",
            type=str,
            default=None,
            help="The name of the project",
        )
        parser.add_argument(
            "-tg",
            "--testmo_gui_url",
            type=str,
            default="https://murrelektronik.testmo.net",
            help="Testmo GUI URL",
        )
        parser.add_argument(
            "-tn", "--testmo_user", type=str, default=None, help="Testmo GUI user name"
        )
        parser.add_argument(
            "-tp",
            "--testmo_password",
            type=str,
            default=None,
            help="Testmo GUI password",
        )
        parser.add_argument(
            "-tu",
            "--testmo_url",
            type=str,
            default="https://murrelektronik.testmo.net/api/v1",
            help="Testmo API URL",
        )
        parser.add_argument(
            "-tt", "--testmo_token", type=str, default=None, help="Testmo API token"
        )
        parser.add_argument(
            "-tf", "--testmo_additional_fields", type=str, default=[], help=""
        )
        parser.add_argument(
            "-cf",
            "--case_filter",
            type=str,
            default="not_retired_or_rejected",
            help=f"Filter to apply - available: {','.join(FilterRegistry.all_filters())} (default)",
        )
        parser.add_argument(
            "-nr",
            "--number_of_runs",
            type=int,
            default=6,
            help="how many of the last runs to check. Use -1 for all runs.",
        )
        args = dict(parser.parse_args().__dict__)

        # event_handler = HandlerChain(InterActiveMode())
        event_handler = [
            ApplicationSetup()
        ]  # as list to enable modification by the handler
        event_handler[0].application_setup(locals())
        event_handler[0].application_started(locals())

        projects = args["project_name"]
        for p in projects:
            args["project_name"] = p
            process_project(event_handler[0], **args)
        sys.exit(0)

    main()
